import openpyxl
import threading
from datetime import datetime
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.scrollview import ScrollView
from kivy.uix.dropdown import DropDown
from kivy.uix.popup import Popup
from kivy.core.window import Window
from kivy.clock import Clock

# Set the window size for a 5-inch display phone (approx 480x800 resolution)
Window.size = (480, 800)

class DataEntryApp(App):
    def build(self):
        # Main layout: BoxLayout with vertical orientation to fill the entire screen
        main_layout = BoxLayout(orientation='vertical', padding=5, spacing=5, size_hint=(1, 1))

        # Create ScrollView to contain inputs
        scrollview = ScrollView(size_hint=(1, 1), do_scroll_y=True)
        form_layout = BoxLayout(orientation='vertical', size_hint_y=None, padding=10, spacing=10)
        form_layout.bind(minimum_height=form_layout.setter('height'))

        # Create the input fields for each data item
        self.inputs = {}
        self.dropdown = None

        self.add_input_field(form_layout, "Medicine Name")
        self.add_input_field(form_layout, "Brand Name")
        self.add_input_field(form_layout, "Price", input_type='number')
        self.add_input_field(form_layout, "No. of Units", input_type='number')
        self.add_input_field(form_layout, "Batch Number")
        self.add_input_field(form_layout, "Supplier Name")

        # Automatically set today's date for the "Date of Purchase"
        self.add_input_field(form_layout, "Date of Purchase", default_value=datetime.today().strftime('%d-%m-%Y'))

        self.add_input_field(form_layout, "Date of Expiry", hint="DD-MM-YYYY")
        self.add_input_field(form_layout, "Quantity Purchased", input_type='number')
        self.add_input_field(form_layout, "Quantity Available", input_type='number')

        # Bind event for medicine name autofill and suggestions
        self.inputs["Medicine Name"].bind(text=self.show_suggestions)

        # Add form layout to scrollview
        scrollview.add_widget(form_layout)
        main_layout.add_widget(scrollview)

        # Add Save Button
        save_button = Button(text="Save to Excel", size_hint=(1, None), height=50)
        save_button.bind(on_press=self.start_save_to_excel)
        main_layout.add_widget(save_button)

        return main_layout

    def add_input_field(self, layout, label_text, input_type='text', hint=None, default_value=""):
        """Helper function to add a label and text input field to the form."""
        label = Label(text=label_text, size_hint_y=None, height=40)
        layout.add_widget(label)

        # Use an empty string if the hint is None
        hint = hint if hint is not None else ""

        if input_type == 'number':
            input_field = TextInput(multiline=False, input_filter='int', size_hint_y=None, height=40, text=default_value)
        else:
            input_field = TextInput(multiline=False, size_hint_y=None, height=40, hint_text=hint, text=default_value)

        layout.add_widget(input_field)
        self.inputs[label_text] = input_field

    def show_suggestions(self, instance, value):
        """Display suggestions in a dropdown as the user types in the Medicine Name field."""
        if self.dropdown:
            self.dropdown.dismiss()  # Close any existing dropdown before showing new suggestions

        self.dropdown = DropDown()
        medicine_name = value.strip().lower()

        # Load the AllMedicineList.xlsx file
        try:
            workbook = openpyxl.load_workbook("AllMedicineList.xlsx")
            sheet = workbook.active
        except FileNotFoundError:
            print("Error: AllMedicineList.xlsx not found.")
            return

        # Search for medicine names that start with the entered letters
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming first row has headers
            if row[0].strip().lower().startswith(medicine_name):
                btn = Button(text=row[0], size_hint_y=None, height=44)
                btn.bind(on_release=lambda btn: self.select_medicine_name(btn.text))
                self.dropdown.add_widget(btn)

        # Show the dropdown if there are suggestions
        if len(self.dropdown.container.children) > 0:
            self.dropdown.open(instance)

    def select_medicine_name(self, medicine_name):
        """Set the selected medicine name and autofill the other fields."""
        self.inputs["Medicine Name"].text = medicine_name
        self.dropdown.dismiss()
        self.autofill_medicine_details()

    def autofill_medicine_details(self):
        """Autofill Brand Name, Price, and No. of Units based on the selected Medicine Name."""
        medicine_name = self.inputs["Medicine Name"].text.strip().lower()

        # Load the AllMedicineList.xlsx file
        try:
            workbook = openpyxl.load_workbook("AllMedicineList.xlsx")
            sheet = workbook.active
        except FileNotFoundError:
            print("Error: AllMedicineList.xlsx not found.")
            return

        # Search for the medicine name in the Excel file and autofill the fields
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming first row has headers
            if row[0].strip().lower() == medicine_name:  # row[0] is the medicine name
                self.inputs["Brand Name"].text = row[1]  # row[1] is the brand name
                self.inputs["Price"].text = str(row[2])  # row[2] is the price
                self.inputs["No. of Units"].text = str(row[3])  # row[3] is the number of units
                return  # Stop searching after finding the first match

    def start_save_to_excel(self, instance):
        """Start the process of saving data in a separate thread."""
        save_thread = threading.Thread(target=self.save_to_excel)
        save_thread.start()

    def save_to_excel(self):
        """Save the entered data into an Excel file on a separate thread."""
        # Read the entered data
        data = {key: input_field.text for key, input_field in self.inputs.items()}

        # Perform Excel write operations in a thread-safe manner
        try:
            workbook = openpyxl.load_workbook("medicine_data.xlsx")
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            # Write headers
            headers = ["Medicine Name", "Brand Name", "Price", "No. of Units", "Batch Number",
                       "Supplier Name", "Date of Purchase", "Date of Expiry", "Quantity Purchased",
                       "Quantity Available"]
            sheet.append(headers)

        # Append new data row
        row_data = [
            data["Medicine Name"],
            data["Brand Name"],
            data["Price"],
            data["No. of Units"],
            data["Batch Number"],
            data["Supplier Name"],
            data["Date of Purchase"],
            data["Date of Expiry"],
            data["Quantity Purchased"],
            data["Quantity Available"]
        ]
        sheet.append(row_data)

        # Save the workbook
        workbook.save("medicine_data.xlsx")

        # Use Clock to schedule updating UI on the main thread
        Clock.schedule_once(lambda dt: self.clear_input_fields())
        Clock.schedule_once(lambda dt: self.show_confirmation())

    def clear_input_fields(self):
        """Clear the input fields after saving."""
        for input_field in self.inputs.values():
            input_field.text = ""

    def show_confirmation(self):
        """Display a confirmation message after saving."""
        confirmation_popup = Popup(title='Success',
                                   content=Label(text="Data saved to Excel successfully!"),
                                   size_hint=(None, None), size=(300, 200))
        confirmation_popup.open()

if __name__ == '__main__':
    DataEntryApp().run()
